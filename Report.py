import matplotlib.pyplot as plt
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from jinja2 import FileSystemLoader, Environment
import pdfkit
from StatisticsByCities import StatisticsByCities
from StatisticsByYear import StatisticsByYear


class Report:
    """Библиотека генерации файлов отчёта в виде .pdf .png .xlsx"""

    def generate_excel(self, name_find_vac: str, stat_by_year: StatisticsByYear, stat_by_cities: StatisticsByCities):
        """Генерация XLSX файла отчёта

        :param name_find_vac: Имя запрашиваемой вакансии
        :param stat_by_cities: Класс статистики для генерации таблицы в .xlsx
        :return:
        """
        wb = Workbook()
        ws1 = wb.create_sheet("Статистика по годам")
        ws2 = wb.create_sheet("Статистика по городам")
        del wb['Sheet']
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ft = Font(bold=True)

        ws1.cell(1, 1, 'Год').font = ft
        ws1.cell(1, 2, 'Средняя зарплата').font = ft
        ws1.cell(1, 3, f'Средняя зарплата - {name_find_vac}').font = ft
        ws1.cell(1, 4, 'Количество вакансий').font = ft
        ws1.cell(1, 5, f'Количество вакансий - {name_find_vac}').font = ft
        for index in range(5):
            ws1.cell(1, index + 1).border = thin_border

        self.__add_cell(ws1, stat_by_year.dict_dynamics_slr, 1, True, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_slr, 2, False, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_slr_name, 3, False, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_count_vac, 4, False, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_count_vac_name, 5, False, thin_border, 0)

        ws2.cell(1, 1, 'Город').font = ft
        ws2.cell(1, 2, 'Уровень зарплат').font = ft
        ws2.cell(1, 4, 'Город').font = ft
        ws2.cell(1, 5, 'Доля вакансий').font = ft
        for index in range(5):
            ws2.cell(1, index + 1 if index == (2 or 3) else index + 2).border = thin_border

        self.__add_cell(ws2, stat_by_cities.dict_dynamics_slr_cities, 1, True, thin_border, 10)
        self.__add_cell(ws2, stat_by_cities.dict_dynamics_slr_cities, 2, False, thin_border, 10)
        self.__add_cell(ws2, stat_by_cities.dict_dynamics_count_vac_big_cities, 4, True, thin_border, 10)
        self.__add_cell(ws2, stat_by_cities.dict_dynamics_count_vac_big_cities, 5, False, thin_border, 10)

        self.__auto_width(ws1)
        self.__auto_width(ws2)
        wb.save('Report/report.xlsx')

    def __add_cell(self, ws, data: dict, y: int, key: bool, border: Border, limit: int):
        """Добавдение ячейки

        :param ws: WorkSheet
        :param data: Целевой словарь
        :param y: Координата смещения
        :param key: Ключ вывода ключей True если надо вывести ключи словаря, False если значения
        :param border: Обводка
        :param limit: Ограничение
        :return:
        """
        if key:
            for index, year in enumerate(data.keys()):
                if index + 1 != limit:
                    ws.cell(index + 2, y, year).border = border
                else:
                    break
        elif list(data.values())[0] < 1:
            for index, val in enumerate(data.values()):
                if index + 1 != limit:
                    ws.cell(index + 2, y, str(round(val * 100, 2)) + '%').border = border
                else:
                    break
        else:
            for index, val in enumerate(data.values()):
                if index + 1 != limit:
                    ws.cell(index + 2, y, val).border = border
                else:
                    break

    def __auto_width(self, ws):
        """Функция автоматического выставления границ столбцов

        :param ws: WorkSheet
        :return:
        """
        for column_cells in ws.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length * 1.23

    def generate_image(self, name_find_vac: str, stat_by_year: StatisticsByYear, stat_by_cities: StatisticsByCities):
        """Функция генерации графиков отчета в .png

        :param name_find_vac: Имя запрашиваемой вакансии
        :param stat_by_cities: Класс статистики для генерации графиков
        :return:
        """
        fig = plt.figure(figsize=(10, 6))
        plt.rcParams['font.size'] = '8'
        width = 0.4
        years = np.arange(len(stat_by_year.dict_dynamics_slr.keys()))
        ax = fig.add_subplot(221)
        ax.bar(years - width / 2,
               stat_by_year.dict_dynamics_slr.values(),
               width,
               label='средняя з/п')
        ax.bar(years + width / 2,
               stat_by_year.dict_dynamics_slr_name.values(),
               width,
               label=f'з/п {name_find_vac}')
        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(years)
        ax.set_xticklabels(stat_by_year.dict_dynamics_slr.keys())
        ax.legend()

        bx = fig.add_subplot(222)
        bx.bar(years - width / 2,
               stat_by_year.dict_dynamics_count_vac.values(),
               width,
               label='Количество вакансий')
        bx.bar(years + width / 2,
               stat_by_year.dict_dynamics_count_vac_name.values(),
               width,
               label=f'Количество вакансий\n{name_find_vac}')
        bx.set_title('Количество вакансий по годам')
        bx.set_xticks(years)
        bx.set_xticklabels(stat_by_year.dict_dynamics_slr.keys())
        bx.legend()
        bx.grid(axis='y')

        dynamics_slr_cities_rev = dict(reversed(list(stat_by_cities.dict_dynamics_slr_cities.items())[:10]))
        cities_slr = np.arange(len(dynamics_slr_cities_rev.keys()))
        cx = fig.add_subplot(223)
        cx.barh(cities_slr - width / 2, dynamics_slr_cities_rev.values(), width + 0.2)
        cx.set_title('Уровень зарплат по годам')
        cx.set_yticks(cities_slr)
        cx.set_yticklabels(dynamics_slr_cities_rev.keys())
        cx.grid(axis='x')

        dx = fig.add_subplot(224)
        dynamics_count_vac_cit_rev = dict(reversed(list(stat_by_cities.dict_dynamics_count_vac_all_cities.items())))
        dx.pie(dynamics_count_vac_cit_rev.values(),
               labels=dynamics_count_vac_cit_rev.keys())
        dx.set_title('Доля вакансий по городам')
        dx.axis("equal")
        fig.savefig('Report/graph.png')

    def generate_pdf(self, name_find_vac: str, stat_by_year: StatisticsByYear, stat_by_cities: StatisticsByCities):
        """Функция генерации отчета в виде .pdf совмещающего и графики, и таблицы

        :param name_find_vac: Имя запрашиваемой вакансии
        :param stat_by_cities: Класс статистики для генерации графиков и таблиц
        :return:
        """
        loader = FileSystemLoader('Resources/temp.html')
        env = Environment(loader=loader)
        template = env.get_template('')
        slr_count_vac_sheet = template.render(name=name_find_vac,
                                              year=list(stat_by_year.dict_dynamics_slr),
                                              slr=list(stat_by_year.dict_dynamics_slr.values()),
                                              slr_name=list(stat_by_year.dict_dynamics_slr_name.values()),
                                              count_vac=list(stat_by_year.dict_dynamics_count_vac.values()),
                                              count_vac_name=list(stat_by_year.dict_dynamics_count_vac_name.values()),
                                              city1=list(stat_by_cities.dict_dynamics_slr_cities.keys())[:10],
                                              slr_lvl=list(stat_by_cities.dict_dynamics_slr_cities.values())[:10],
                                              city2=list(stat_by_cities.dict_dynamics_count_vac_big_cities.keys())[:10],
                                              part_slr=list(map(lambda x: str(round(x * 100, 4)) + '%',
                                                                stat_by_cities.dict_dynamics_count_vac_big_cities.values()))[
                                                       :10])
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(slr_count_vac_sheet,
                           'Report/report.pdf',
                           configuration=config,
                           options={'enable-local-file-access': None})