import csv
import math
import os
import re
from collections import Counter
from itertools import islice
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from jinja2 import FileSystemLoader, Environment
import pdfkit
from prettytable import PrettyTable, ALL
from multiprocessing import Manager, Pool
import cProfile, pstats, io
from pstats import SortKey
from concurrent.futures import ProcessPoolExecutor
from dateutil import parser
from datetime import datetime

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

all_titles_table = ["№", "Название", "Оклад", "Название региона", "Дата публикации вакансии"]

dict_experince_format = {
    "noExperience": "Нет опыта",
    "between1And3": "От 1 года до 3 лет",
    "between3And6": "От 3 до 6 лет",
    "moreThan6": "Более 6 лет"
}

dict_slr_currency = {
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум"
}


class Vacancy:
    """Класс вакансии

    Attributes:
        name (str): Имя
        salary (Salary): Зарплата
        area_name (str): Регион
        published_at (str): Дата публикации
        published_at_year (str): Год публикации
    """

    def __init__(self, dict_vacancy: dict):
        """Инициализирует класс вакансии из словаря вакансии

        :param dict_vacancy: Словарь вакансии
        """
        self.name = dict_vacancy['name']
        self.salary = Salary(salary_from=dict_vacancy['salary_from'],
                             salary_to=dict_vacancy['salary_to'],
                             salary_currency=dict_vacancy['salary_currency'])
        self.area_name = dict_vacancy['area_name']
        self.published_at = dict_vacancy['published_at']

        # str(parser.parse(dict_vacancy['published_at']).date())
        # '.'.join(str(datetime.datetime.strptime(dict_vacancy['published_at'], '%Y-%m-%dT%H:%M:%S%z').date()).split('-'))

        self.published_at_year = int(self.published_at[:4])


class Salary:
    """Класс зарплаты
    Attributes:
        salary_from (int): Оклад 'от'
        salary_to (int): Оклад 'до'
        salary_currency (str): Валюта
    """

    def __init__(self, salary_from, salary_to, salary_currency: str):
        """Инициализация класса зарплаты

        :param salary_from(int, float, str): Оклад 'от'
        :param salary_to(int, float, str): Оклад 'до'
        :param salary_currency: Валюта
        >>> type(Salary(10, 20, 'RUR')).__name__
        'Salary'
        >>> Salary(10.0, 20.0, 'RUR').salary_from
        10
        >>> Salary(10.0, 20.0, 'RUR').salary_to
        20
        >>> Salary(10.0, 20.0, 'RUR').salary_currency
        'RUR'
        """
        self.salary_from = int(float(salary_from))
        self.salary_to = int(float(salary_to))
        if salary_currency not in currency_to_rub.keys():
            raise ValueError('Неверно введна валюта')
        self.salary_currency = salary_currency

    def get_salary_to_rub(self) -> float:
        """Функция подсчета средней ЗП в рублях с помощью словаря currency_to_rub
        :return: Вывод средней ЗП в рублях
        >>> Salary(20.0, 30.0, 'RUR').get_salary_to_rub()
        25.0
        >>> Salary(20.0, 30.0, 'RUR').salary_from
        20
        >>> Salary(20.0, 30.0, 'RUR').salary_to
        30
        >>> Salary(20.0, 30.0, 'EUR').get_salary_to_rub()
        1497.5
        >>> Salary(20, 30.0, 'RUR').get_salary_to_rub()
        25.0
        """
        return currency_to_rub[self.salary_currency] * (float(self.salary_to) + float(self.salary_from)) / 2


class DataSet:
    """Класс датасета

    Attributes:
        file_name (str): Имя файла
        vacancies_objects (list): Список вакансий
        vacancies_object_name (list): Список вакансий по заданному имени
    """

    def __init__(self, file_name: str, name: str, start: int, end: int):
        """Инициализация класса датасета

        :param file_name: Имя файла
        :param name: Имя искомой профессии
        :param start: С какого года выводить информацию
        :param end: По какой год выводить информацию
        >>> DataSet('Tests/test_data_set.csv', 'аналитик', 2007, 2014).file_name
        'Tests/test_data_set.csv'
        >>> DataSet('Tests/test_data_set.csv', 'администратор', 2007, 2014).vacancies_objects[0].name
        'Менеджер по работе с юридическими лицами'
        >>> DataSet('Tests/test_data_set.csv', 'администратор', 2007, 2014).vacancies_objects_name[0].name
        'Системный администратор'
        """
        self.file_name = file_name
        self.vacancies_objects = []
        self.vacancies_objects_name = []
        for vacancy in self.__csv_reader():
            if not start <= int(vacancy['published_at'][:4]) <= end:
                continue
            self.vacancies_objects.append(Vacancy(vacancy))
            if name in vacancy['name']:
                self.vacancies_objects_name.append(Vacancy(vacancy))

    def __csv_reader(self) -> list:
        """Чтение .csv

        :return: Вывод списка обработанных данных
        """
        with open(self.file_name, encoding='utf-8-sig') as read_file:
            file_reader = csv.reader(read_file, delimiter=",")
            data_vacancies = []
            try:
                titles = next(file_reader)
            except StopIteration:
                print('Пустой файл')
                exit(0)
            html_tags = re.compile('<.*?>')
            for row in file_reader:
                if '' in row or len(row) != len(titles):
                    continue
                self.__csv_filer(html_tags=html_tags,
                                 row=row,
                                 titles=titles)
                if len(row) == len(titles):
                    data_vacancies.append({titles[i]: row[i] for i in range(len(titles))})
            if len(data_vacancies) == 0:
                print('Нет данных')
                exit(0)
            return data_vacancies

    @staticmethod
    def __csv_filer(html_tags, row: list, titles: list):
        """Проверка строки вакансии на правильность и отсев 'неправильных'

        :param html_tags: html теги
        :param row: Строка вакансии
        :param titles: Заголовки
        :return:
        """
        if len(row) < len(titles):
            return
        for i in range(len(row)):
            row[i] = ' '.join(re.sub(html_tags, '', row[i])
                              .strip()
                              .split())


class InputConnect:
    """Класс ввода с консоли и вывода таблицы в консоль

    Attributes:
        file_name (str): Имя файла
        name (str): Искомое имя
        __dict_formatter (dict): Словарь функций предобработки вакансий для вывода таблицы
    """

    def __init__(self):
        """Инициализация класса InputConnect"""
        self.file_name = self.__processing_file_name(input('Введите название файла: '))
        self.name = input('Введите название профессии: ')
        self.__dict_formatter = {
            'Название': lambda row: row.name,
            'Оклад': lambda
                row: f'{self.__slr_format(row.salary.salary_from)} - {self.__slr_format(row.salary.salary_to)} ' \
                     f'({dict_slr_currency[row.salary.salary_currency]}) ',
            'Название региона': lambda row: row.area_name,
            'Дата публикации вакансии': lambda row: '.'.join(reversed(row.published_at[0:10].split('-')))
        }

    @staticmethod
    def __processing_file_name(file_name: str) -> str:
        """Обработка ввода имени файла

        :param file_name: Имя файла
        :return: Имя файла
        """
        if os.stat(file_name).st_size == 0:
            print('Пустой файл')
            exit(0)
        return file_name

    @staticmethod
    def __slr_format(slr: str) -> str:
        """Форматирование ЗП для данных таблицы

        :param slr: Зарплата 'от' или 'до'
        :return: Преобразованная ЗП
        """
        return '{:,}'.format(math.floor(float(slr))).replace(',', ' ')

    def table_print(self, data_vacancies: list):
        """Вывод в консоль таблицы

        :param data_vacancies: Список вакансий
        :return:
        """
        counter = 1
        table = PrettyTable()
        table.hrules = ALL
        table.align = 'l'
        table.field_names = all_titles_table
        for title in table.field_names:
            table.max_width[title] = 20
        for value in data_vacancies:
            temp_array = [counter]
            for v in all_titles_table[1:]:
                temp = self.__dict_formatter[v](value)
                temp_array.append(temp if len(temp) < 100 else temp[:100] + '...')
            table.add_row(temp_array)
            counter += 1
        print(table.get_string(fields=all_titles_table))

    def split(self, delimiter=',', output_path='csv_files_by_year'):
        with open(self.file_name, encoding='utf-8-sig') as read_file:
            file_reader = csv.reader(read_file, delimiter=delimiter)
            try:
                titles = next(file_reader)
            except StopIteration:
                print('Пустой файл')
                exit(0)

            first_row = next(file_reader)
            current_year = first_row[titles.index('published_at')][2:4]
            current_out_path = output_path + '/' + current_year + '.csv'
            file_names = [current_out_path]
            current_out_writer = csv.writer(open(current_out_path, 'w', encoding='utf-8-sig'),
                                            delimiter=delimiter,
                                            lineterminator="\r")
            current_out_writer.writerow(titles)
            current_out_writer.writerow(first_row)
            for row in file_reader:
                if row[titles.index('published_at')][2:4] != current_year:
                    current_year = row[titles.index('published_at')][2:4]
                    current_out_path = output_path + '/' + current_year + '.csv'
                    file_names.append(current_out_path)
                    current_out_writer = csv.writer(open(current_out_path, 'w', encoding='utf-8-sig'),
                                                    delimiter=delimiter,
                                                    lineterminator="\r")
                    current_out_writer.writerow(titles)
                    current_out_writer.writerow(row)
                else:
                    current_out_writer.writerow(row)
            return file_names


class StatisticsByYear:
    def __init__(self, name: str, splitted_file_names: list):
        self.name = name
        self.list_dict_dynamics_slr = []
        self.list_dict_dynamics_count_vac = []
        self.list_dict_dynamics_slr_name = []
        self.list_dict_dynamics_count_vac_name = []

        pool = ProcessPoolExecutor(max_workers=6)
        for el in pool.map(self.stat, splitted_file_names):
            self.list_dict_dynamics_slr.append(el[0])
            self.list_dict_dynamics_count_vac.append(el[1])
            self.list_dict_dynamics_slr_name.append(el[2])
            self.list_dict_dynamics_count_vac_name.append(el[3])

        self.list_dict_dynamics_slr = list(filter(None, self.list_dict_dynamics_slr))
        self.list_dict_dynamics_count_vac = list(filter(None, self.list_dict_dynamics_count_vac))
        self.list_dict_dynamics_slr_name = list(filter(None, self.list_dict_dynamics_slr_name))
        self.list_dict_dynamics_count_vac_name = list(filter(None, self.list_dict_dynamics_count_vac_name))

        self.dict_dynamics_slr = self.__get_dict_from_list(self.list_dict_dynamics_slr)
        self.dict_dynamics_count_vac = self.__get_dict_from_list(self.list_dict_dynamics_count_vac)
        self.dict_dynamics_slr_name = self.__get_dict_from_list(self.list_dict_dynamics_slr_name)
        self.dict_dynamics_count_vac_name = self.__get_dict_from_list(self.list_dict_dynamics_count_vac_name)

    @staticmethod
    def __get_dict_from_list(list_dicts: list) -> dict:
        result = {}
        for x in list_dicts:
            result.update(x)
        return result

    def stat(self, file_name):
        dataset = DataSet(file_name, self.name, 2007, 2014)
        return self.__dynamics_salary(dataset.vacancies_objects), \
               self.__dynamics_count_vac(dataset.vacancies_objects), \
               self.__dynamics_salary(dataset.vacancies_objects_name), \
               self.__dynamics_count_vac(dataset.vacancies_objects_name)

    def __dynamics_salary(self, data_vacancies: list) -> dict:
        """Составление словаря динамики уровня зарплат по годам для всех вакансий

        :param data_vacancies: Список вакансий
        :return: Словарь динамики уровня зарплат по годам для всех вакансий
        """
        # if len(data_vacancies) == 0:
        #    return {x: 0 for x in self.__list_years}
        dict_dynamic_slr = {}
        for vac in data_vacancies:
            if vac.published_at_year not in dict_dynamic_slr.keys():
                dict_dynamic_slr[vac.published_at_year] = [vac.salary.get_salary_to_rub()]
            else:
                dict_dynamic_slr[vac.published_at_year].append(vac.salary.get_salary_to_rub())
        for key, value in dict_dynamic_slr.items():
            dict_dynamic_slr[key] = math.floor(sum(value) / len(value))
        return dict_dynamic_slr

    def __dynamics_count_vac(self, data_vacancies: list) -> dict:
        """Составление словаря динамики количества вакансий по годам для всех вакансий

        :param data_vacancies: Список вакансий
        :return: Словарь динамики количества вакансий по годам для всех вакансий
        """
        # if len(data_vacancies) == 0:
        #    return {x: 0 for x in self.__list_years}
        dict_dynamic_count_vac = dict(Counter(map(lambda x: x.published_at_year, data_vacancies)))
        return dict_dynamic_count_vac

    def print_statistics(self):
        """Метод печати статистики в консоль

        :return:
        """
        print('Динамика уровня зарплат по годам: ' + str(self.dict_dynamics_slr))
        print('Динамика количества вакансий по годам: ' + str(self.dict_dynamics_count_vac))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(self.dict_dynamics_slr_name))
        print(
            'Динамика количества вакансий по годам для выбранной профессии: ' + str(self.dict_dynamics_count_vac_name))


class StatisticsByCities:
    """Класс статистики

    Attributes:
        dict_dynamics_count_vac_all_cities (dict): Словарь количества вакансий по всем городам
        dict_dynamics_count_vac_big_cities (dict): Словарь количества вакансий по 'большим' городам
        dict_dynamics_slr_cities (dict): Словарь уровня зарплат по 'большим' городам
    """

    def __init__(self, file_name: str, name: str):
        """Инициадизация класса Statistics

        :param dataset: Объект класса DataSet
        """
        dataset = DataSet(file_name, name, 2007, 2014)
        self.dict_dynamics_count_vac_all_cities = self.__dynamics_count_vac_cities(dataset.vacancies_objects)
        self.dict_dynamics_count_vac_big_cities = dict(filter(lambda x: x[0] != 'Другие',
                                                              list(self.dict_dynamics_count_vac_all_cities.items())))
        self.dict_dynamics_slr_cities = self.__dynamics_slr_big_cities(data_vacancies=dataset.vacancies_objects,
                                                                       big_cities=list(
                                                                           self.dict_dynamics_count_vac_big_cities))

    def __dynamics_count_vac_cities(self, data_vacancies: list) -> dict:
        """Составление словаря количества вакансий по городам

        :param data_vacancies: Список вакансий
        :return: Словарь количества вакансий по городам
        """
        dict_dynamic_count_vac_cities = dict(Counter(map(lambda x: x.area_name, data_vacancies)))
        dict_dynamic_count_vac_big_cities = {}
        count_vac_small_cities = 0
        for key, val in dict_dynamic_count_vac_cities.items():
            if math.floor(val * 100 / len(data_vacancies)) >= 1:
                dict_dynamic_count_vac_big_cities[key] = round(val / len(data_vacancies), 4)
            else:
                count_vac_small_cities += round(val / len(data_vacancies), 4)
        dict_dynamic_count_vac_big_cities['Другие'] = count_vac_small_cities
        return dict(sorted(dict_dynamic_count_vac_big_cities.items(), key=lambda x: x[1], reverse=True))

    @staticmethod
    def __dynamics_slr_big_cities(data_vacancies: list, big_cities: list) -> dict:
        """Составление словаря уровня зарплат по 'большим' городам

        :param data_vacancies: Список вакансий
        :param big_cities: Список 'больших' городов
        :return: Словарь уровня зарплат по 'большим' городам
        """
        dict_dynamic_slr_cities = {}
        for vac in data_vacancies:
            if vac.area_name not in dict_dynamic_slr_cities.keys() and vac.area_name in big_cities:
                dict_dynamic_slr_cities[vac.area_name] = [vac.salary.get_salary_to_rub()]
            elif vac.area_name in big_cities:
                dict_dynamic_slr_cities[vac.area_name].append(vac.salary.get_salary_to_rub())
        for key, value in dict_dynamic_slr_cities.items():
            dict_dynamic_slr_cities[key] = math.floor(sum(value) / len(value))
        return dict(sorted(dict_dynamic_slr_cities.items(), key=lambda x: x[1], reverse=True))

    def print_statistics(self):
        """Метод печати статистики в консоль

        :return:
        """
        print('Уровень зарплат по городам (в порядке убывания): ' +
              str(dict(islice(self.dict_dynamics_slr_cities.items(), 10))))
        print('Доля вакансий по городам (в порядке убывания): ' +
              str(dict(islice(self.dict_dynamics_count_vac_big_cities.items(), 10))))


class Report:
    """Библиотека генерации файлов отчёта в виде .pdf .png .xlsx"""

    def generate_excel(self, name_find_vac: str, stat_by_year: StatisticsByYear, stat_by_cities: StatisticsByCities):
        """Генерация XLSX файла отчёта

        :param name_find_vac: Имя запрашиваемой вакансии
        :param stat_by_cities: Класс статистики для генерации таблицы в .xlsx
        :return:
        """
        wb = Workbook()
        ws1 = wb.create_sheet("Статистика по годам")
        ws2 = wb.create_sheet("Статистика по городам")
        del wb['Sheet']
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ft = Font(bold=True)

        ws1.cell(1, 1, 'Год').font = ft
        ws1.cell(1, 2, 'Средняя зарплата').font = ft
        ws1.cell(1, 3, f'Средняя зарплата - {name_find_vac}').font = ft
        ws1.cell(1, 4, 'Количество вакансий').font = ft
        ws1.cell(1, 5, f'Количество вакансий - {name_find_vac}').font = ft
        for index in range(5):
            ws1.cell(1, index + 1).border = thin_border

        self.__add_cell(ws1, stat_by_year.dict_dynamics_slr, 1, True, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_slr, 2, False, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_slr_name, 3, False, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_count_vac, 4, False, thin_border, 0)
        self.__add_cell(ws1, stat_by_year.dict_dynamics_count_vac_name, 5, False, thin_border, 0)

        ws2.cell(1, 1, 'Город').font = ft
        ws2.cell(1, 2, 'Уровень зарплат').font = ft
        ws2.cell(1, 4, 'Город').font = ft
        ws2.cell(1, 5, 'Доля вакансий').font = ft
        for index in range(5):
            ws2.cell(1, index + 1 if index == (2 or 3) else index + 2).border = thin_border

        self.__add_cell(ws2, stat_by_cities.dict_dynamics_slr_cities, 1, True, thin_border, 10)
        self.__add_cell(ws2, stat_by_cities.dict_dynamics_slr_cities, 2, False, thin_border, 10)
        self.__add_cell(ws2, stat_by_cities.dict_dynamics_count_vac_big_cities, 4, True, thin_border, 10)
        self.__add_cell(ws2, stat_by_cities.dict_dynamics_count_vac_big_cities, 5, False, thin_border, 10)

        self.__auto_width(ws1)
        self.__auto_width(ws2)
        wb.save('report.xlsx')

    def __add_cell(self, ws, data: dict, y: int, key: bool, border: Border, limit: int):
        """Добавдение ячейки

        :param ws: WorkSheet
        :param data: Целевой словарь
        :param y: Координата смещения
        :param key: Ключ вывода ключей True если надо вывести ключи словаря, False если значения
        :param border: Обводка
        :param limit: Ограничение
        :return:
        """
        if key:
            for index, year in enumerate(data.keys()):
                if index + 1 != limit:
                    ws.cell(index + 2, y, year).border = border
                else:
                    break
        elif list(data.values())[0] < 1:
            for index, val in enumerate(data.values()):
                if index + 1 != limit:
                    ws.cell(index + 2, y, str(round(val * 100, 2)) + '%').border = border
                else:
                    break
        else:
            for index, val in enumerate(data.values()):
                if index + 1 != limit:
                    ws.cell(index + 2, y, val).border = border
                else:
                    break

    def __auto_width(self, ws):
        """Функция автоматического выставления границ столбцов

        :param ws: WorkSheet
        :return:
        """
        for column_cells in ws.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length * 1.23

    def generate_image(self, name_find_vac: str, stat_by_year: StatisticsByYear, stat_by_cities: StatisticsByCities):
        """Функция генерации графиков отчета в .png

        :param name_find_vac: Имя запрашиваемой вакансии
        :param stat_by_cities: Класс статистики для генерации графиков
        :return:
        """
        fig = plt.figure(figsize=(10, 6))
        plt.rcParams['font.size'] = '8'
        width = 0.4
        years = np.arange(len(stat_by_year.dict_dynamics_slr.keys()))
        ax = fig.add_subplot(221)
        ax.bar(years - width / 2,
               stat_by_year.dict_dynamics_slr.values(),
               width,
               label='средняя з/п')
        ax.bar(years + width / 2,
               stat_by_year.dict_dynamics_slr_name.values(),
               width,
               label=f'з/п {name_find_vac}')
        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(years)
        ax.set_xticklabels(stat_by_year.dict_dynamics_slr.keys())
        ax.legend()

        bx = fig.add_subplot(222)
        bx.bar(years - width / 2,
               stat_by_year.dict_dynamics_count_vac.values(),
               width,
               label='Количество вакансий')
        bx.bar(years + width / 2,
               stat_by_year.dict_dynamics_count_vac_name.values(),
               width,
               label=f'Количество вакансий\n{name_find_vac}')
        bx.set_title('Количество вакансий по годам')
        bx.set_xticks(years)
        bx.set_xticklabels(stat_by_year.dict_dynamics_slr.keys())
        bx.legend()
        bx.grid(axis='y')

        dynamics_slr_cities_rev = dict(reversed(list(stat_by_cities.dict_dynamics_slr_cities.items())[:10]))
        cities_slr = np.arange(len(dynamics_slr_cities_rev.keys()))
        cx = fig.add_subplot(223)
        cx.barh(cities_slr - width / 2, dynamics_slr_cities_rev.values(), width + 0.2)
        cx.set_title('Уровень зарплат по годам')
        cx.set_yticks(cities_slr)
        cx.set_yticklabels(dynamics_slr_cities_rev.keys())
        cx.grid(axis='x')

        dx = fig.add_subplot(224)
        dynamics_count_vac_cit_rev = dict(reversed(list(stat_by_cities.dict_dynamics_count_vac_all_cities.items())))
        dx.pie(dynamics_count_vac_cit_rev.values(),
               labels=dynamics_count_vac_cit_rev.keys())
        dx.set_title('Доля вакансий по городам')
        dx.axis("equal")
        fig.savefig('graph.png')

    def generate_pdf(self, name_find_vac: str, stat_by_year: StatisticsByYear, stat_by_cities: StatisticsByCities):
        """Функция генерации отчета в виде .pdf совмещающего и графики, и таблицы

        :param name_find_vac: Имя запрашиваемой вакансии
        :param stat_by_cities: Класс статистики для генерации графиков и таблиц
        :return:
        """
        loader = FileSystemLoader('./temp.html')
        env = Environment(loader=loader)
        template = env.get_template('')
        slr_count_vac_sheet = template.render(name=name_find_vac,
                                              year=list(stat_by_year.dict_dynamics_slr),
                                              slr=list(stat_by_year.dict_dynamics_slr.values()),
                                              slr_name=list(stat_by_year.dict_dynamics_slr_name.values()),
                                              count_vac=list(stat_by_year.dict_dynamics_count_vac.values()),
                                              count_vac_name=list(stat_by_year.dict_dynamics_count_vac_name.values()),
                                              city1=list(stat_by_cities.dict_dynamics_slr_cities.keys())[:10],
                                              slr_lvl=list(stat_by_cities.dict_dynamics_slr_cities.values())[:10],
                                              city2=list(stat_by_cities.dict_dynamics_count_vac_big_cities.keys())[:10],
                                              part_slr=list(map(lambda x: str(round(x * 100, 4)) + '%',
                                                                stat_by_cities.dict_dynamics_count_vac_big_cities.values()))[
                                                       :10])
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(slr_count_vac_sheet,
                           'report.pdf',
                           configuration=config,
                           options={'enable-local-file-access': None})


if __name__ == '__main__':
    pr = cProfile.Profile()
    pr.enable()
    input_data = InputConnect()
    changing_output = int(input('Таблица в консоль или отчет по статистике? (1 или 2): '))
    if changing_output == 2:
        dataset = DataSet(input_data.file_name, input_data.name, 2007, 2014)
        splitted_file_names = input_data.split()
        statistics_by_year = StatisticsByYear(input_data.name, splitted_file_names)
        statistics_by_cities = StatisticsByCities(input_data.file_name, input_data.name)
        statistics_by_cities.print_statistics()
        statistics_by_year.print_statistics()
        report = Report()
        report.generate_excel(input_data.name, statistics_by_year, statistics_by_cities)
        report.generate_image(input_data.name, statistics_by_year, statistics_by_cities)
        report.generate_pdf(input_data.name, statistics_by_year, statistics_by_cities)
    elif changing_output == 1:
        dataset = DataSet(input_data.file_name, input_data.name, 2007, 2014)
        input_data.table_print(dataset.vacancies_objects_name)
    pr.disable()
    s = io.StringIO()
    sortby = SortKey.CUMULATIVE
    ps = pstats.Stats(pr, stream=s).sort_stats(sortby)
    ps.print_stats()
    print(s.getvalue())
